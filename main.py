from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
import os
from datetime import datetime

app = FastAPI(title="JSON to Excel Converter", version="1.0.0")

class ExcelData(BaseModel):
    data: List[Dict[str, Any]]

# Жёстко заданная последовательность колонок (как на скрине)
PREFERRED_ORDER = [
    "Category",
    "Brand",
    "Model",
    "Color",
    "DDP (RUB)",
    "Valid Until",
    "Markup Serb",
    "Markup RU",
    "Exchange Rate",
    "Qty Offered",
    "Date Offered",
    "Qty Ordered",
    "Date Ordered",
]

def create_formatted_excel(data: List[Dict[str, Any]]) -> str:
    """
    Создает красиво отформатированный Excel файл из JSON данных,
    фиксируя порядок колонок по PREFERRED_ORDER.
    Поля, которых нет в PREFERRED_ORDER, добавляются в конец
    в порядке первого появления в данных.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Export"

    if not data:
        return None

    # 1) Берём из предпочтительного списка только те поля, которые реально встречаются
    headers: List[str] = []
    seen = set()
    for k in PREFERRED_ORDER:
        if any(k in row for row in data):
            headers.append(k)
            seen.add(k)

    # 2) Добавляем любые другие поля, встреченные в данных, в порядке первого появления
    for row in data:
        for k in row.keys():
            if k not in seen:
                headers.append(k)
                seen.add(k)

    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    data_font = Font(size=11)
    data_alignment = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Данные (None -> "", чтобы не было "None" в Excel)
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            if value is None:
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    # Автоподбор ширины
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_len = len(str(headers[col-1]))
        for row in range(2, len(data) + 2):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[column_letter].width = min(max_len + 2, 50)

    ws.freeze_panes = "A2"

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file.close()
    wb.save(temp_file.name)
    return temp_file.name

@app.post("/convert-to-excel")
async def convert_to_excel(excel_data: ExcelData):
    try:
        if not excel_data.data:
            raise HTTPException(status_code=400, detail="Данные не могут быть пустыми")
        excel_file_path = create_formatted_excel(excel_data.data)
        if not excel_file_path:
            raise HTTPException(status_code=500, detail="Ошибка при создании Excel файла")
        return FileResponse(
            path=excel_file_path,
            filename=f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка при обработке данных: {str(e)}")

@app.get("/")
async def root():
    return {
        "message": "JSON to Excel Converter API",
        "version": "1.0.0",
        "endpoints": {
            "POST /convert-to-excel": "Конвертирует JSON данные в Excel файл",
            "GET /": "Информация о API"
        }
    }

if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    host = os.getenv("HOST", "0.0.0.0")
    uvicorn.run(app, host=host, port=port)
